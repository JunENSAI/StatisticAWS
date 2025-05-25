import json
from urllib.parse import unquote_plus
import boto3
from botocore.exceptions import ClientError
import os
import logging
import csv
import io
# openpyxl devra être inclus dans le package de déploiement de la Lambda
try:
    import openpyxl
except ImportError:
    openpyxl = None # Gérer le cas où il n'est pas disponible, bien que ce soit une dépendance clé


logger = logging.getLogger()
logger.setLevel("INFO")

s3_client = boto3.client('s3')
dynamodb_resource = None
files_table = None # Renommé pour correspondre à notre cas d'usage

# Utiliser la variable d'environnement pour la table des fichiers
FILES_DYNAMO_TABLE_NAME = os.getenv("DYNAMO_TABLE", "MyFilesTable") # Correspond à ce qui est dans app_files.py

if FILES_DYNAMO_TABLE_NAME:
    try:
        dynamodb_resource = boto3.resource('dynamodb')
        files_table = dynamodb_resource.Table(FILES_DYNAMO_TABLE_NAME)
        logger.info(f"Successfully initialized DynamoDB table object for table: {FILES_DYNAMO_TABLE_NAME}")
    except Exception as e:
        logger.error(f"Failed to initialize DynamoDB table resource for table name '{FILES_DYNAMO_TABLE_NAME}': {e}", exc_info=True)
else:
    logger.error("Environment variable FILES_DYNAMO_TABLE is not set!")


def extract_csv_metadata(file_content_stream):
    try:
        content = file_content_stream.read().decode('utf-8-sig') # utf-8-sig gère le BOM
        # Essayer de détecter le délimiteur ou supposer le point-virgule si c'est courant pour vos fichiers
        # Option 1: Utiliser csv.Sniffer pour détecter le dialecte (plus robuste mais peut échouer)
        try:
            dialect = csv.Sniffer().sniff(content.splitlines()[0]) # Sniff sur la première ligne
            logger.info(f"CSV dialect sniffed: delimiter='{dialect.delimiter}', quotechar='{dialect.quotechar}'")
            reader = csv.reader(io.StringIO(content), dialect=dialect)
        except csv.Error:
            logger.warning("CSV Sniffer failed, falling back to ';' delimiter.")
            # Option 2: Supposer le point-virgule si le sniff échoue ou si vous savez que c'est le cas
            reader = csv.reader(io.StringIO(content), delimiter=';') 

        rows = list(reader)
        if not rows:
            logger.warning("CSV file appears to be empty or unparseable with current delimiter.")
            return None, 0, 0
        
        headers = rows[0]
        num_rows = len(rows) - 1
        num_cols = len(headers)
        
        logger.info(f"Extracted headers: {headers}, Num_cols: {num_cols}, Num_rows: {num_rows}")
        if num_cols <= 1 and ';' in content.splitlines()[0]: # Si on a une seule colonne mais qu'il y a des ';' dans l'en-tête
            logger.warning("Possible delimiter issue: Only one column detected but ';' present in header. Check delimiter.")

        return headers, num_rows, num_cols
    except Exception as e:
        logger.error(f"Error processing CSV content: {e}", exc_info=True)
        raise


def extract_excel_metadata(file_content_stream):
    """Extrait les métadonnées d'un fichier Excel (.xlsx)."""
    if not openpyxl:
        logger.error("openpyxl library is not available. Cannot process Excel files.")
        raise ImportError("openpyxl library not found")
    try:
        workbook = openpyxl.load_workbook(filename=file_content_stream) # Pas besoin de io.BytesIO si openpyxl > 2.5
        sheet = workbook.active # Prendre la première feuille active
        
        if sheet.max_row == 0: # Feuille vide
             return None, 0, 0

        headers = [cell.value for cell in sheet[1]] # Première ligne pour les en-têtes
        num_rows = sheet.max_row - 1 # Exclure la ligne d'en-tête
        num_cols = sheet.max_column

        # S'assurer que les en-têtes vides à la fin sont retirés si num_cols est basé sur la cellule la plus à droite
        actual_headers = [h for h in headers if h is not None]
        if len(actual_headers) < num_cols and headers: # si il y a des headers mais moins que max_cols
             if all(h is None for h in headers[len(actual_headers):]):
                 num_cols = len(actual_headers)
                 headers = actual_headers
        elif not headers and num_cols > 0: # Pas de headers mais des colonnes ?
            headers = [f"Column_{i+1}" for i in range(num_cols)]


        return headers, num_rows, num_cols
    except Exception as e:
        logger.error(f"Error processing Excel content: {e}", exc_info=True)
        raise


def lambda_handler(event, context):
    if not files_table:
        logger.error("DynamoDB 'files_table' resource is not initialized. Aborting.")
        return {'statusCode': 500, 'body': json.dumps('Internal server error: Files table not configured or initialization failed')}

    for record in event.get("Records", []):
        processing_status = "processed_with_metadata" # Statut par défaut
        extracted_metadata = {}

        try:
            s3_data = record.get("s3", {})
            bucket_name = s3_data.get("bucket", {}).get("name")
            object_key = s3_data.get("object", {}).get("key")

            if not bucket_name or not object_key:
                logger.warning(f"Skipping record due to missing bucket name or object key: {record}")
                continue

            # Décoder la clé de l'objet (peut contenir des caractères spéciaux comme '+')
            key = unquote_plus(object_key)
            logger.info(f"Processing object s3://{bucket_name}/{key}")

            # Structure de clé attendue : user_uploads/{user}/{file_id}/{s3_filename}
            parts = key.split('/')
            if len(parts) < 4 or parts[0] != "user_uploads":
                logger.error(f"Invalid key format: '{key}'. Expected 'user_uploads/user/file_id/filename'. Skipping.")
                continue

            user = parts[1]
            file_id = parts[2]
            # s3_filename = parts[3] # Peut être utile pour le logging

            logger.info(f"Extracted from key: user='{user}', file_id='{file_id}'")

            # Télécharger le fichier depuis S3
            try:
                s3_object = s3_client.get_object(Bucket=bucket_name, Key=key)
                file_content_stream = s3_object['Body'] # Ceci est un flux
            except ClientError as e:
                logger.error(f"S3 GetObject error for key '{key}': {e}", exc_info=True)
                processing_status = "error_s3_read"
                continue # Passer à l'enregistrement suivant

            # Déterminer le type de fichier et extraire les métadonnées
            headers = None
            num_rows = 0
            num_cols = 0

            try:
                if key.lower().endswith('.csv'):
                    logger.info(f"Processing as CSV: {key}")
                    headers, num_rows, num_cols = extract_csv_metadata(file_content_stream)
                elif key.lower().endswith('.xlsx'):
                    if not openpyxl:
                         logger.error("openpyxl not available, cannot process .xlsx file.")
                         processing_status = "error_missing_dependency_xlsx"
                         raise RuntimeError("openpyxl not available")
                    logger.info(f"Processing as Excel (xlsx): {key}")
                    # openpyxl attend un objet de type fichier binaire pour les flux
                    headers, num_rows, num_cols = extract_excel_metadata(io.BytesIO(file_content_stream.read()))
                else:
                    logger.warning(f"Unsupported file type for key: {key}. Skipping metadata extraction.")
                    processing_status = "unsupported_file_type"
                    # Pas besoin de 'continue' ici si on veut quand même mettre à jour DynamoDB avec ce statut
                
                if headers is not None: # Si le parsing a réussi
                    extracted_metadata = {
                        'columnHeaders': headers,
                        'rowCount': num_rows,
                        'columnCount': num_cols
                    }
                    logger.info(f"Extracted metadata for {key}: Rows={num_rows}, Cols={num_cols}, Headers={headers[:5]}...") # Log seulement les premiers headers

            except Exception as e: # Erreur pendant le parsing du fichier
                logger.error(f"Failed to parse file content for {key}: {e}", exc_info=True)
                processing_status = "error_parsing_file"
                # On continue pour mettre à jour DynamoDB avec ce statut d'erreur

            # Mettre à jour l'item dans DynamoDB
            update_expression_parts = ["SET processingStatus = :ps"]
            expression_attribute_values = {':ps': processing_status}
            
            if extracted_metadata: # N'ajouter que si on a des métadonnées
                update_expression_parts.append("columnHeaders = :ch")
                expression_attribute_values[':ch'] = extracted_metadata.get('columnHeaders', [])
                update_expression_parts.append("rowCount = :rc")
                expression_attribute_values[':rc'] = extracted_metadata.get('rowCount', 0)
                update_expression_parts.append("columnCount = :cc")
                expression_attribute_values[':cc'] = extracted_metadata.get('columnCount', 0)
                update_expression_parts.append("processedTimestamp = :pt") # Ajouter un timestamp de traitement
                expression_attribute_values[':pt'] = datetime.datetime.utcnow().isoformat()


            update_expression = ", ".join(update_expression_parts)

            logger.info(f"Attempting to update DynamoDB item with Key: user='{user}', file_id='{file_id}'")
            logger.debug(f"UpdateExpression: {update_expression}")
            logger.debug(f"ExpressionAttributeValues: {expression_attribute_values}")

            try:
                update_response = files_table.update_item(
                    Key={
                        'user': user,
                        'file_id': file_id # Notre clé de tri pour la table des fichiers
                    },
                    UpdateExpression=update_expression,
                    ExpressionAttributeValues=expression_attribute_values,
                    ReturnValues="UPDATED_NEW"
                )
                logger.info(f"DynamoDB update successful for file '{file_id}'. Updated attributes: {update_response.get('Attributes')}")

            except ClientError as e:
                if e.response['Error']['Code'] == 'ConditionalCheckFailedException': # Ne devrait pas arriver sans ConditionExpression
                    logger.error(f"DynamoDB update failed for file '{file_id}': Item does not exist.", exc_info=True)
                else:
                    logger.error(f"DynamoDB ClientError updating file '{file_id}': {e}", exc_info=True)
                # Si la mise à jour DynamoDB échoue, l'erreur est loggée, on passe au record suivant.
                continue 
            except Exception as e:
                logger.error(f"Unexpected error updating DynamoDB for file '{file_id}': {e}", exc_info=True)
                continue

        except Exception as e: # Erreur générale de traitement d'un record
            logger.error(f"Error processing record: {record}. Error: {e}", exc_info=True)
            # Pourrait envoyer à une Dead Letter Queue (DLQ) si configuré
            continue

    return {
        'statusCode': 200,
        'body': json.dumps('Finished processing S3 event for file metadata.')
    }
